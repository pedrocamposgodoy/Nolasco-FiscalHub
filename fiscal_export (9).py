# ================================================================
# fiscal_export.py — Módulo de Exportación Fiscal para Asesor
# Nolasco Capital · Módulo independiente
#
# Contiene:
#   1. calcular_resumen_global()  — suma todos los inmuebles → RESULTADOS
#   2. generar_excel_asesor()     — Excel formato Álvaro con logo
#   3. generar_pdf_global()       — PDF multi-inmueble con todos sumados
#   4. render_seccion_fiscal()    — Streamlit: pantalla Fiscalidad completa
#
# Uso en app.py:
#   from fiscal_export import render_seccion_fiscal
#   elif menu == "Fiscalidad":
#       render_seccion_fiscal(df_inm, df_mov, safe_float, calcular_modelo_100)
# ================================================================

import io
import pandas as pd
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.platypus import Table, TableStyle
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False


# ────────────────────────────────────────────────────────────────
# 1. CALCULAR RESUMEN GLOBAL — agrega todos los inmuebles
# ────────────────────────────────────────────────────────────────

def calcular_resumen_global(df_inm, df_mov, safe_float_fn, calcular_modelo_100_fn,
                             año_fiscal=2025):
    """
    Calcula el Modelo 100 para cada inmueble y devuelve:
      - lista de dicts por inmueble con todas las casillas
      - dict de totales agregados (la hoja RESULTADOS del Excel de Álvaro)
    """
    filas = []
    for _, row in df_inm.iterrows():
        modelo = calcular_modelo_100_fn(row, df_mov, año_fiscal=año_fiscal)
        filas.append({
            "inmueble":          row.get("Nombre", ""),
            "ref_catastral":     row.get("Ref_Catastral", "N/A"),
            "tipo":              row.get("Tipo_Arrendamiento", "Larga Duración"),
            "inquilino":         row.get("Inquilino", ""),
            "nif_inquilino":     row.get("NIF_Inquilino", ""),
            "dias":              modelo["0101"],
            "ingresos":          modelo["0102"],
            "intereses":         modelo["0105"],
            "reparaciones":      modelo["0106"],
            "ibi":               modelo["0108"],
            "comunidad_seguros": modelo["0110"],
            "suministros":       modelo["0111"],
            "gastos_juridicos":  modelo["0112"],
            "amortizacion":      modelo["0113"],
            "amort_detalle":     modelo.get("0113_detalle", ""),
            "total_gastos":      modelo["0107"],
            "rend_neto":         modelo["0149"],
            "reduccion_pct":     modelo["reduccion_pct"],
            "reduccion_imp":     modelo["0150"],
            "rend_final":        modelo["0152"],
            "retenciones":       modelo["0153"],
            "nota_reduccion":    modelo.get("nota_reduccion", ""),
        })

    # Totales
    totales = {
        "ingresos":          sum(f["ingresos"]          for f in filas),
        "intereses":         sum(f["intereses"]         for f in filas),
        "reparaciones":      sum(f["reparaciones"]      for f in filas),
        "ibi":               sum(f["ibi"]               for f in filas),
        "comunidad_seguros": sum(f["comunidad_seguros"] for f in filas),
        "suministros":       sum(f["suministros"]       for f in filas),
        "gastos_juridicos":  sum(f["gastos_juridicos"]  for f in filas),
        "amortizacion":      sum(f["amortizacion"]      for f in filas),
        "total_gastos":      sum(f["total_gastos"]      for f in filas),
        "rend_neto":         sum(f["rend_neto"]         for f in filas),
        "reduccion_imp":     sum(f["reduccion_imp"]     for f in filas),
        "rend_final":        sum(f["rend_final"]        for f in filas),
        "retenciones":       sum(f["retenciones"]       for f in filas),
        "n_inmuebles":       len(filas),
        "año_fiscal":        año_fiscal,
    }

    return filas, totales


# ────────────────────────────────────────────────────────────────
# 2. EXPORT EXCEL — formato Álvaro con logo asesoría
# ────────────────────────────────────────────────────────────────

def generar_excel_asesor(filas, totales, nombre_propietario="Propietario",
                          nombre_asesoria="", año_fiscal=2025):
    """
    Genera un Excel con 3 hojas:
      - CONTABILIDAD : una columna por inmueble (formato Álvaro)
      - RESULTADOS   : totales agregados
      - BORRADOR FISCAL : casillas Modelo 100 con advertencias
    Devuelve BytesIO listo para st.download_button
    """
    if not OPENPYXL_OK:
        return None

    wb = openpyxl.Workbook()

    # ── Colores ──────────────────────────────────────────────────
    AZUL_OSC  = "0F2744"
    AZUL_ACC  = "185FA5"
    AZUL_CLAR = "D0DFF0"
    GRIS_CLAR = "F4F7FB"
    AMARILLO  = "FFF9E6"
    VERDE     = "E8F5E9"
    ROJO_CLAR = "FFEBEE"
    BLANCO    = "FFFFFF"

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size)

    def _border():
        thin = Side(style="thin", color="D0DFF0")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _fmt(ws, row, col, value, bold=False, bg=None, fg="000000",
             align="left", num_fmt=None, size=10, wrap=False, border=True):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = _font(bold=bold, color=fg, size=size)
        cell.alignment = _aln(h=align, wrap=wrap)
        if bg:
            cell.fill = _fill(bg)
        if num_fmt:
            cell.number_format = num_fmt
        if border:
            cell.border = _border()
        return cell

    # ═══════════════════════════════════════════════════════════
    # HOJA 1 — CONTABILIDAD (una columna por inmueble)
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "CONTABILIDAD"
    ws1.sheet_view.showGridLines = False

    # Cabecera corporativa
    ws1.row_dimensions[1].height = 35
    ws1.row_dimensions[2].height = 18
    n_cols = len(filas) + 2  # col A = concepto, cols B... = inmuebles, última = TOTAL

    ws1.merge_cells(start_row=1, start_column=1,
                    end_row=1, end_column=n_cols + 1)
    c = ws1.cell(row=1, column=1,
                 value=f"NOLASCO CAPITAL — Rendimientos Capital Inmobiliario {año_fiscal}")
    c.font      = Font(bold=True, color=BLANCO, size=13)
    c.fill      = _fill(AZUL_OSC)
    c.alignment = _aln(h="center")

    ws1.merge_cells(start_row=2, start_column=1,
                    end_row=2, end_column=n_cols + 1)
    c2 = ws1.cell(row=2, column=1,
                  value=f"Propietario: {nombre_propietario}   |   "
                        f"Asesoría: {nombre_asesoria}   |   "
                        f"Generado: {datetime.now().strftime('%d/%m/%Y')}")
    c2.font      = Font(italic=True, color="5A7A9A", size=9)
    c2.fill      = _fill(GRIS_CLAR)
    c2.alignment = _aln(h="center")

    # Fila 3 — cabeceras de inmuebles
    ws1.row_dimensions[3].height = 30
    _fmt(ws1, 3, 1, "CONCEPTO", bold=True, bg=AZUL_OSC, fg=BLANCO,
         align="center", size=9)

    for i, f in enumerate(filas):
        col = i + 2
        c = _fmt(ws1, 3, col, f["inmueble"], bold=True, bg=AZUL_OSC,
                 fg=BLANCO, align="center", size=9, wrap=True)
        ws1.column_dimensions[get_column_letter(col)].width = 18

    _fmt(ws1, 3, len(filas) + 2, "TOTALES", bold=True, bg=AZUL_ACC,
         fg=BLANCO, align="center", size=9)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions[get_column_letter(len(filas) + 2)].width = 14

    # Filas de datos
    CONCEPTOS = [
        # (label, campo, es_seccion, formato, bg_override)
        ("DATOS DEL INMUEBLE",      None,              True,  None,    AZUL_ACC),
        ("Ref. Catastral",           "ref_catastral",   False, None,    None),
        ("Inquilino",                "inquilino",        False, None,    None),
        ("NIF Inquilino",            "nif_inquilino",    False, None,    None),
        ("Tipo Arrendamiento",       "tipo",             False, None,    None),
        ("Días arrendados (0101)",   "dias",             False, "0",     None),
        ("INGRESOS",                 None,               True,  None,    AZUL_ACC),
        ("Ingresos íntegros (0102)", "ingresos",         False, "#,##0.00 €", VERDE),
        ("GASTOS DEDUCIBLES",        None,               True,  None,    AZUL_ACC),
        ("Intereses hipoteca (0105)","intereses",        False, "#,##0.00 €", None),
        ("Reparación/conserv. (0106)","reparaciones",   False, "#,##0.00 €", None),
        ("IBI y tributos (0108)",    "ibi",              False, "#,##0.00 €", None),
        ("Comunidad+Seguros (0110)", "comunidad_seguros",False, "#,##0.00 €", None),
        ("Suministros (0111)",       "suministros",      False, "#,##0.00 €", None),
        ("Gastos jurídicos (0112)",  "gastos_juridicos", False, "#,##0.00 €", None),
        ("Amortización 3% (0113)",   "amortizacion",     False, "#,##0.00 €", None),
        ("TOTAL GASTOS (0107)",      "total_gastos",     False, "#,##0.00 €", ROJO_CLAR),
        ("RESULTADOS",               None,               True,  None,    AZUL_ACC),
        ("Rendimiento Neto (0149)",  "rend_neto",        False, "#,##0.00 €", GRIS_CLAR),
        ("Reducción % aplicable",    "reduccion_pct",    False, '0"%"',  AMARILLO),
        ("Reducción importe (0150)", "reduccion_imp",    False, "#,##0.00 €", AMARILLO),
        ("Rend. Neto Reducido (0152)","rend_final",      False, "#,##0.00 €", VERDE),
        ("Retenciones (0153)",       "retenciones",      False, "#,##0.00 €", None),
    ]

    fila_actual = 4
    for concepto, campo, es_seccion, fmt, bg_override in CONCEPTOS:
        ws1.row_dimensions[fila_actual].height = 22 if es_seccion else 18
        bg = bg_override or (GRIS_CLAR if fila_actual % 2 == 0 else BLANCO)

        if es_seccion:
            # Fila de sección — fusionar y centrar
            ws1.merge_cells(start_row=fila_actual, start_column=1,
                             end_row=fila_actual, end_column=len(filas) + 2)
            _fmt(ws1, fila_actual, 1, concepto, bold=True, bg=bg,
                 fg=BLANCO, align="center", size=9)
        else:
            _fmt(ws1, fila_actual, 1, concepto, bold=False, bg=GRIS_CLAR,
                 fg="0F2744", align="left", size=9)
            # Valores por inmueble
            for i, f in enumerate(filas):
                val = f.get(campo, "") if campo else ""
                col = i + 2
                bg_cel = bg_override or (GRIS_CLAR if fila_actual % 2 == 0 else BLANCO)
                c = _fmt(ws1, fila_actual, col, val, bold=False, bg=bg_cel,
                          fg="0D1B2A", align="right", size=9)
                if fmt and val != "":
                    c.number_format = fmt

            # Columna TOTAL (solo para campos numéricos)
            if campo and campo not in ("ref_catastral", "inquilino",
                                        "nif_inquilino", "tipo",
                                        "amort_detalle", "nota_reduccion"):
                total_val = totales.get(campo, "")
                col_tot = len(filas) + 2
                bg_tot = bg_override or BLANCO
                ct = _fmt(ws1, fila_actual, col_tot, total_val,
                           bold=True, bg=bg_tot, fg="0F2744",
                           align="right", size=9)
                if fmt and total_val != "":
                    ct.number_format = fmt

        fila_actual += 1

    # Nota advertencia reducción
    fila_actual += 1
    ws1.merge_cells(start_row=fila_actual, start_column=1,
                    end_row=fila_actual, end_column=len(filas) + 2)
    nota = ws1.cell(row=fila_actual, column=1,
                    value="⚠️  REDUCCIÓN: Orientativa. Contratos anteriores 26/05/2023 → 60%. "
                           "Posteriores → 50% general (verificar 60/70/90% según condiciones). "
                           "Validar siempre con asesor fiscal.")
    nota.font      = Font(italic=True, color="854F0B", size=8)
    nota.fill      = _fill(AMARILLO)
    nota.alignment = _aln(h="left", wrap=True)
    ws1.row_dimensions[fila_actual].height = 30

    # Congelar cabeceras
    ws1.freeze_panes = "B4"

    # ═══════════════════════════════════════════════════════════
    # HOJA 2 — RESULTADOS (formato hoja RESULTADOS de Álvaro)
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("RESULTADOS")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 18

    # Header
    ws2.merge_cells("A1:B1")
    c = ws2.cell(row=1, column=1,
                 value=f"RESULTADOS — IRPF {año_fiscal}")
    c.font = Font(bold=True, color=BLANCO, size=13)
    c.fill = _fill(AZUL_OSC)
    c.alignment = _aln(h="center")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:B2")
    c2 = ws2.cell(row=2, column=1,
                  value=f"Propietario: {nombre_propietario}  |  "
                        f"{totales['n_inmuebles']} inmuebles  |  "
                        f"Generado: {datetime.now().strftime('%d/%m/%Y')}")
    c2.font = Font(italic=True, color="5A7A9A", size=9)
    c2.fill = _fill(GRIS_CLAR)
    c2.alignment = _aln(h="center")
    ws2.row_dimensions[2].height = 18

    filas_resultado = [
        ("INGRESOS",                          None,                   True,  AZUL_ACC),
        ("Ingresos íntegros totales",          "ingresos",             False, VERDE),
        ("GASTOS",                             None,                   True,  AZUL_ACC),
        ("Intereses hipoteca",                 "intereses",            False, None),
        ("Reparación y conservación",          "reparaciones",         False, None),
        ("IBI y tributos",                     "ibi",                  False, None),
        ("Comunidad + Seguros",                "comunidad_seguros",    False, None),
        ("Servicios y suministros",            "suministros",          False, None),
        ("Gastos jurídicos",                   "gastos_juridicos",     False, None),
        ("Amortización fiscal (3%)",           "amortizacion",         False, None),
        ("TOTAL GASTOS DEDUCIBLES",            "total_gastos",         False, ROJO_CLAR),
        ("RENDIMIENTO NETO (antes reducción)", "rend_neto",            False, GRIS_CLAR),
        ("REDUCCIÓN FISCAL",                   None,                   True,  AZUL_ACC),
        ("Reducción aplicable (orientativa)",  "reduccion_imp",        False, AMARILLO),
        ("RENDIMIENTO NETO REDUCIDO",          "rend_final",           False, VERDE),
        ("Retenciones practicadas",            "retenciones",          False, None),
    ]

    fila_r = 3
    for label, campo, es_sec, bg_ov in filas_resultado:
        ws2.row_dimensions[fila_r].height = 24 if es_sec else 19
        bg = bg_ov or (GRIS_CLAR if fila_r % 2 == 0 else BLANCO)

        _fmt(ws2, fila_r, 1, label,
             bold=es_sec, bg=bg, fg=BLANCO if es_sec else "0F2744",
             align="left" if not es_sec else "center", size=10)

        if not es_sec and campo:
            val = totales.get(campo, 0)
            c = _fmt(ws2, fila_r, 2, val,
                     bold=(bg_ov in (VERDE, ROJO_CLAR)),
                     bg=bg, fg="0F2744", align="right", size=10)
            c.number_format = "#,##0.00 €"

        elif es_sec:
            ws2.merge_cells(
                start_row=fila_r, start_column=1,
                end_row=fila_r, end_column=2
            )

        fila_r += 1

    # Nota reducción en RESULTADOS
    fila_r += 1
    ws2.merge_cells(start_row=fila_r, start_column=1,
                    end_row=fila_r, end_column=2)
    nota2 = ws2.cell(row=fila_r, column=1,
                     value="⚠️  Reducción orientativa. Validar porcentaje exacto con asesor fiscal "
                            "según fecha contrato e ingresos totales del contribuyente.")
    nota2.font      = Font(italic=True, color="854F0B", size=8)
    nota2.fill      = _fill(AMARILLO)
    nota2.alignment = _aln(h="left", wrap=True)
    ws2.row_dimensions[fila_r].height = 30

    # ═══════════════════════════════════════════════════════════
    # HOJA 3 — BORRADOR FISCAL (casillas Modelo 100)
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("BORRADOR FISCAL")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 40

    # Header
    ws3.merge_cells("A1:D1")
    c = ws3.cell(row=1, column=1,
                 value=f"BORRADOR MODELO 100 — Casillas IRPF {año_fiscal}")
    c.font = Font(bold=True, color=BLANCO, size=13)
    c.fill = _fill(AZUL_OSC)
    c.alignment = _aln(h="center")
    ws3.row_dimensions[1].height = 30

    ws3.merge_cells("A2:D2")
    c2 = ws3.cell(row=2, column=1,
                  value=f"Propietario: {nombre_propietario}  |  "
                        f"Asesoría: {nombre_asesoria}  |  "
                        f"Datos orientativos — pendiente validación asesor")
    c2.font = Font(italic=True, color="854F0B", size=9)
    c2.fill = _fill(AMARILLO)
    c2.alignment = _aln(h="center")
    ws3.row_dimensions[2].height = 18

    # Cabeceras tabla
    for col, label in enumerate(["Casilla", "Descripción", "Importe Total", "Nota para asesor"], 1):
        _fmt(ws3, 3, col, label, bold=True, bg=AZUL_OSC, fg=BLANCO,
             align="center", size=9)
    ws3.row_dimensions[3].height = 22

    casillas_borrador = [
        ("0101", "Días arrendados (promedio)",           f"{365} días",
         "Verificar días exactos por inmueble"),
        ("0102", "Ingresos íntegros totales",            totales["ingresos"],
         ""),
        ("0105", "Intereses hipoteca/financiación",      totales["intereses"],
         "Solo intereses, no amortización capital"),
        ("0106", "Reparación y conservación",            totales["reparaciones"],
         "Del diario contable Nolasco Capital"),
        ("0108", "Tributos e IBI",                       totales["ibi"],
         ""),
        ("0110", "Comunidad + Seguros + Ascensor",       totales["comunidad_seguros"],
         "Seguro hogar + vida + comunidad"),
        ("0111", "Servicios y suministros",              totales["suministros"],
         ""),
        ("0112", "Gastos jurídicos y administrativos",   totales["gastos_juridicos"],
         ""),
        ("0113", "Amortización fiscal (3%)",             totales["amortizacion"],
         "MAX(precio compra, catastral) × %construcción × 3%"),
        ("0107", "TOTAL GASTOS DEDUCIBLES",              totales["total_gastos"],
         ""),
        ("0149", "RENDIMIENTO NETO",                     totales["rend_neto"],
         ""),
        ("0150", "Reducción (orientativa)",              totales["reduccion_imp"],
         "⚠️ VALIDAR: 60% ant. 26/05/23 · 50% post · verificar 70/90%"),
        ("0152", "RENDIMIENTO NETO REDUCIDO",            totales["rend_final"],
         "⚠️ VALIDAR con asesor fiscal"),
        ("0153", "Retenciones practicadas",              totales["retenciones"],
         ""),
    ]

    fila_b = 4
    for casilla, desc, val, nota in casillas_borrador:
        ws3.row_dimensions[fila_b].height = 19
        es_total = casilla in ("0107", "0149", "0152")
        bg = VERDE if casilla == "0152" else (ROJO_CLAR if casilla == "0107" else
              AMARILLO if casilla in ("0150",) else
              GRIS_CLAR if fila_b % 2 == 0 else BLANCO)

        _fmt(ws3, fila_b, 1, casilla, bold=es_total, bg=bg, fg="0F2744",
             align="center", size=9)
        _fmt(ws3, fila_b, 2, desc, bold=es_total, bg=bg, fg="0F2744",
             align="left", size=9)

        c = _fmt(ws3, fila_b, 3, val if isinstance(val, str) else float(val),
                 bold=es_total, bg=bg, fg="0F2744", align="right", size=9)
        if not isinstance(val, str):
            c.number_format = "#,##0.00 €"

        nota_color = "854F0B" if "⚠️" in nota else "5A7A9A"
        cn = _fmt(ws3, fila_b, 4, nota.replace("⚠️ ", ""), bold=False, bg=bg,
                  fg=nota_color, align="left", size=8, wrap=True)

        fila_b += 1

    # Footer legal
    fila_b += 1
    ws3.merge_cells(start_row=fila_b, start_column=1,
                    end_row=fila_b, end_column=4)
    footer = ws3.cell(row=fila_b, column=1,
                      value="IMPORTANTE: Documento informativo generado por Nolasco Capital. "
                             "No sustituye el asesoramiento fiscal profesional. "
                             "Amortización: MAX(precio compra total, catastral) × % construcción × 3% × % titularidad (Art. 14 RIRPF). "
                             f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    footer.font      = Font(italic=True, color="5A7A9A", size=7.5)
    footer.fill      = _fill(GRIS_CLAR)
    footer.alignment = _aln(h="left", wrap=True)
    ws3.row_dimensions[fila_b].height = 35

    # Guardar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ────────────────────────────────────────────────────────────────
# 3. PDF GLOBAL — todos los inmuebles + resumen final
# ────────────────────────────────────────────────────────────────

def generar_pdf_global(filas, totales, nombre_propietario="Propietario",
                        nombre_asesoria="", año_fiscal=2025):
    """
    Genera PDF multi-página:
      - Portada
      - Una página por inmueble con sus casillas
      - Página final: RESUMEN GLOBAL (todos sumados)
    Devuelve BytesIO.
    """
    if not REPORTLAB_OK:
        return None

    buffer = io.BytesIO()
    c = rl_canvas.Canvas(buffer, pagesize=A4)
    w, h = A4

    AZ_OSC  = HexColor("#0F2744")
    AZ_ACC  = HexColor("#185FA5")
    GRIS    = HexColor("#F4F7FB")
    BORDE   = HexColor("#D0DFF0")
    VERDE_C = HexColor("#E8F5E9")
    VERDE_T = HexColor("#1a7a40")
    NARANJA = HexColor("#B45309")
    AMAR    = HexColor("#FFF9E6")
    ROJO_C  = HexColor("#FFEBEE")
    ref     = f"NC-{año_fiscal}-{nombre_propietario[:3].upper()}"

    def _header(titulo_extra=""):
        c.setFillColor(AZ_OSC)
        c.rect(0, h - 85, w, 85, fill=True, stroke=False)
        c.setFillColor(AZ_ACC)
        c.roundRect(25, h - 72, 50, 50, 5, fill=True, stroke=False)
        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 18)
        c.drawCentredString(50, h - 52, "NC")
        c.setFont("Helvetica", 7)
        c.drawCentredString(50, h - 64, "CAPITAL")
        c.setFont("Helvetica-Bold", 17)
        c.drawString(90, h - 43, "Nolasco Capital")
        c.setFont("Helvetica", 8)
        c.drawString(90, h - 57, "GRANADA  |  GESTIÓN PATRIMONIAL INMOBILIARIA")
        if titulo_extra:
            c.setFont("Helvetica-Oblique", 8)
            c.drawString(90, h - 70, titulo_extra)
        c.setFont("Helvetica", 7)
        c.drawRightString(w - 25, h - 40, f"Ref: {ref}")
        c.drawRightString(w - 25, h - 52, f"Ejercicio: {año_fiscal}")
        c.drawRightString(w - 25, h - 64, f"{datetime.now().strftime('%d/%m/%Y')}")
        if nombre_asesoria:
            c.setFillColor(AZ_ACC)
            c.roundRect(w - 160, h - 78, 135, 13, 3, fill=True, stroke=False)
            c.setFillColor(white)
            c.setFont("Helvetica-Bold", 7)
            c.drawCentredString(w - 92, h - 70, nombre_asesoria[:35])
        c.setStrokeColor(AZ_ACC)
        c.setLineWidth(2.5)
        c.line(0, h - 87, w, h - 87)

    def _footer():
        c.setFillColor(AZ_OSC)
        c.rect(0, 0, w, 25, fill=True, stroke=False)
        c.setFillColor(white)
        c.setFont("Helvetica", 6.5)
        c.drawString(25, 9, "Nolasco Capital  |  Documento informativo — pendiente validación asesor fiscal")
        c.drawRightString(w - 25, 9, f"Propietario: {nombre_propietario}  |  IRPF {año_fiscal}")

    # ── PORTADA ──────────────────────────────────────────────────
    _header()
    _footer()

    y = h - 140
    c.setFillColor(AZ_OSC)
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(w / 2, y, f"Borrador IRPF {año_fiscal}")
    c.setFont("Helvetica", 11)
    c.setFillColor(HexColor("#5A7A9A"))
    c.drawCentredString(w / 2, y - 22, "Rendimientos de Capital Inmobiliario — Modelo 100")

    y -= 55
    c.setFillColor(GRIS)
    c.roundRect(40, y - 80, w - 80, 90, 8, fill=True, stroke=False)
    c.setStrokeColor(BORDE)
    c.roundRect(40, y - 80, w - 80, 90, 8, fill=False, stroke=True)
    c.setFillColor(AZ_OSC)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(55, y - 10, "Datos del contribuyente")
    c.setFont("Helvetica", 9)
    c.setFillColor(HexColor("#5A7A9A"))
    c.drawString(55, y - 26, "Nombre:")
    c.drawString(55, y - 42, "Asesoría fiscal:")
    c.drawString(55, y - 58, "Nº inmuebles:")
    c.drawString(55, y - 74, "Generado por:")
    c.setFillColor(AZ_OSC)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(165, y - 26, nombre_propietario)
    c.drawString(165, y - 42, nombre_asesoria or "—")
    c.drawString(165, y - 58, str(totales["n_inmuebles"]))
    c.drawString(165, y - 74, f"Nolasco Capital  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    # KPIs portada
    y -= 110
    kpis = [
        ("Ingresos totales",     totales["ingresos"],    AZ_ACC),
        ("Total gastos",         totales["total_gastos"], HexColor("#C0392B")),
        ("Rend. Neto",           totales["rend_neto"],   AZ_OSC),
        ("Base imponible est.",  totales["rend_final"],  VERDE_T),
    ]
    kw = (w - 80) / 4
    for i, (label, val, color) in enumerate(kpis):
        x = 40 + i * kw
        c.setFillColor(color)
        c.roundRect(x, y - 55, kw - 8, 60, 6, fill=True, stroke=False)
        c.setFillColor(white)
        c.setFont("Helvetica", 7)
        c.drawCentredString(x + kw / 2 - 4, y - 10, label.upper())
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(x + kw / 2 - 4, y - 32, f"{val:,.0f}€")
        c.setFont("Helvetica", 7)
        c.drawCentredString(x + kw / 2 - 4, y - 46, "IRPF estimado")

    # Índice inmuebles
    y -= 90
    c.setFillColor(AZ_OSC)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y, "Inmuebles incluidos en este borrador:")
    c.setStrokeColor(AZ_ACC)
    c.setLineWidth(1.5)
    c.line(40, y - 4, 260, y - 4)
    y -= 20
    for i, f in enumerate(filas):
        c.setFillColor(AZ_ACC if i % 2 == 0 else AZ_OSC)
        c.roundRect(40, y - 14, 12, 12, 2, fill=True, stroke=False)
        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(46, y - 8, str(i + 1))
        c.setFillColor(AZ_OSC)
        c.setFont("Helvetica", 8)
        c.drawString(58, y - 8,
                     f"{f['inmueble']}  ·  {f['tipo']}  ·  "
                     f"Rend. neto: {f['rend_neto']:,.0f}€  ·  "
                     f"Amort: {f['amortizacion']:,.0f}€")
        y -= 18

    c.showPage()

    # ── UNA PÁGINA POR INMUEBLE ──────────────────────────────────
    for idx, f in enumerate(filas):
        _header(f"Inmueble {idx + 1} de {len(filas)}: {f['inmueble']}")
        _footer()
        y = h - 115

        # Datos inmueble
        c.setFillColor(GRIS)
        c.roundRect(25, y - 55, w - 50, 58, 6, fill=True, stroke=False)
        c.setStrokeColor(BORDE)
        c.roundRect(25, y - 55, w - 50, 58, 6, fill=False, stroke=True)
        c.setFillColor(AZ_OSC)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(35, y - 12, "Datos del inmueble")
        c.setFont("Helvetica", 8)
        c.setFillColor(HexColor("#5A7A9A"))
        pairs1 = [("Inmueble:", f["inmueble"]),
                  ("Ref. Catastral:", f["ref_catastral"]),
                  ("Titular/NIF:", nombre_propietario)]
        pairs2 = [("Modalidad:", f["tipo"]),
                  ("Inquilino:", f["inquilino"]),
                  ("NIF Inquilino:", f["nif_inquilino"])]
        for j, (lbl, val) in enumerate(pairs1):
            x = 35 + j * 175
            c.drawString(x, y - 26, lbl)
            c.setFillColor(AZ_OSC)
            c.setFont("Helvetica-Bold", 8)
            c.drawString(x + 72, y - 26, str(val)[:22])
            c.setFont("Helvetica", 8)
            c.setFillColor(HexColor("#5A7A9A"))
        for j, (lbl, val) in enumerate(pairs2):
            x = 35 + j * 175
            c.drawString(x, y - 42, lbl)
            c.setFillColor(AZ_OSC)
            c.setFont("Helvetica-Bold", 8)
            c.drawString(x + 72, y - 42, str(val)[:22])
            c.setFont("Helvetica", 8)
            c.setFillColor(HexColor("#5A7A9A"))

        # Tabla casillas
        y -= 75
        c.setFillColor(AZ_OSC)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(25, y, "Casillas Modelo 100")
        c.setStrokeColor(AZ_ACC)
        c.setLineWidth(1.5)
        c.line(25, y - 4, 200, y - 4)
        y -= 18

        tabla_data = [
            ["Casilla", "Descripción", "Importe (€)", "Nota"],
            ["0101", "Días arrendados",             f"{f['dias']} días", ""],
            ["0102", "Ingresos íntegros",            f"{f['ingresos']:,.2f}", ""],
            ["0105", "Intereses hipoteca",           f"{f['intereses']:,.2f}", "Solo intereses"],
            ["0106", "Reparación y conservación",    f"{f['reparaciones']:,.2f}", "Diario contable"],
            ["0108", "IBI y tributos",               f"{f['ibi']:,.2f}", ""],
            ["0110", "Comunidad + Seguros",          f"{f['comunidad_seguros']:,.2f}", "Hogar+vida+comunidad"],
            ["0111", "Suministros",                  f"{f['suministros']:,.2f}", ""],
            ["0112", "Gastos jurídicos",             f"{f['gastos_juridicos']:,.2f}", ""],
            ["0113", "Amortización 3%",              f"{f['amortizacion']:,.2f}",
             f"{f['amort_detalle'][:30]}..." if len(f['amort_detalle']) > 30
             else f['amort_detalle']],
            ["0107", "TOTAL GASTOS",                f"{f['total_gastos']:,.2f}", ""],
            ["0149", "RENDIMIENTO NETO",             f"{f['rend_neto']:,.2f}", ""],
            ["0150", f"Reducción {f['reduccion_pct']}% (orient.)",
             f"-{f['reduccion_imp']:,.2f}", "⚠️ VALIDAR con asesor"],
            ["0152", "REND. NETO REDUCIDO",          f"{f['rend_final']:,.2f}", "⚠️ VALIDAR"],
            ["0153", "Retenciones",                  f"{f['retenciones']:,.2f}", ""],
        ]

        t = Table(tabla_data, colWidths=[48, 165, 85, 145])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0),  (-1, 0),  AZ_OSC),
            ("TEXTCOLOR",     (0, 0),  (-1, 0),  white),
            ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0),  (-1, 0),  8),
            ("FONTNAME",      (0, 1),  (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1),  (-1, -1), 7.5),
            ("ALIGN",         (2, 1),  (2, -1),  "RIGHT"),
            ("FONTNAME",      (0, 1),  (0, -1),  "Helvetica-Bold"),
            ("GRID",          (0, 0),  (-1, -1), 0.4, BORDE),
            ("LINEBELOW",     (0, 0),  (-1, 0),  2, AZ_ACC),
            ("ROWBACKGROUNDS",(0, 1),  (-1, -5), [white, GRIS]),
            ("BACKGROUND",    (0, -4), (-1, -4), GRIS),
            ("FONTNAME",      (0, -4), (-1, -4), "Helvetica-Bold"),
            ("BACKGROUND",    (0, -3), (-1, -3), AMAR),
            ("FONTNAME",      (0, -3), (-1, -3), "Helvetica-Bold"),
            ("TEXTCOLOR",     (3, -3), (3, -3),  NARANJA),
            ("BACKGROUND",    (0, -2), (-1, -2), HexColor("#FFF0DC")),
            ("FONTNAME",      (0, -2), (-1, -2), "Helvetica-Bold"),
            ("TEXTCOLOR",     (3, -2), (3, -2),  NARANJA),
            ("BACKGROUND",    (0, -1), (-1, -1), GRIS),
            ("TOPPADDING",    (0, 0),  (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
            ("LEFTPADDING",   (0, 0),  (-1, -1), 5),
        ]))
        t.wrapOn(c, w, h)
        t.drawOn(c, 22, y - len(tabla_data) * 18 - 5)

        c.showPage()

    # ── PÁGINA FINAL: RESUMEN GLOBAL ─────────────────────────────
    _header("RESUMEN GLOBAL — Todos los inmuebles")
    _footer()
    y = h - 115

    c.setFillColor(AZ_OSC)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(w / 2, y, f"RESUMEN GLOBAL — {totales['n_inmuebles']} INMUEBLES")
    c.setFont("Helvetica", 9)
    c.setFillColor(HexColor("#5A7A9A"))
    c.drawCentredString(w / 2, y - 16,
                        "Suma de todos los rendimientos de capital inmobiliario")
    y -= 40

    resumen_data = [
        ["Casilla", "Concepto", "Importe Total (€)", "Estado"],
        ["0102", "Ingresos íntegros totales",          f"{totales['ingresos']:,.2f}",      "✓"],
        ["0105", "Intereses hipoteca",                  f"{totales['intereses']:,.2f}",     "✓"],
        ["0106", "Reparación y conservación",           f"{totales['reparaciones']:,.2f}",  "✓"],
        ["0108", "IBI y tributos",                      f"{totales['ibi']:,.2f}",            "✓"],
        ["0110", "Comunidad + Seguros",                 f"{totales['comunidad_seguros']:,.2f}", "✓"],
        ["0111", "Suministros",                         f"{totales['suministros']:,.2f}",   "✓"],
        ["0112", "Gastos jurídicos",                    f"{totales['gastos_juridicos']:,.2f}","✓"],
        ["0113", "Amortización fiscal 3%",              f"{totales['amortizacion']:,.2f}",  "✓"],
        ["0107", "TOTAL GASTOS DEDUCIBLES",             f"{totales['total_gastos']:,.2f}",  ""],
        ["0149", "RENDIMIENTO NETO",                    f"{totales['rend_neto']:,.2f}",     ""],
        ["0150", "Reducción fiscal (orientativa)",      f"-{totales['reduccion_imp']:,.2f}","⚠️"],
        ["0152", "RENDIMIENTO NETO REDUCIDO (BASE)",    f"{totales['rend_final']:,.2f}",    "⚠️"],
        ["0153", "Retenciones practicadas",             f"{totales['retenciones']:,.2f}",   "✓"],
    ]

    t2 = Table(resumen_data, colWidths=[48, 190, 110, 40])
    t2.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  AZ_OSC),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  white),
        ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0),  (-1, 0),  9),
        ("FONTNAME",      (0, 1),  (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1),  (-1, -1), 8.5),
        ("ALIGN",         (2, 1),  (2, -1),  "RIGHT"),
        ("ALIGN",         (3, 0),  (3, -1),  "CENTER"),
        ("FONTNAME",      (0, 1),  (0, -1),  "Helvetica-Bold"),
        ("GRID",          (0, 0),  (-1, -1), 0.4, BORDE),
        ("LINEBELOW",     (0, 0),  (-1, 0),  2, AZ_ACC),
        ("ROWBACKGROUNDS",(0, 1),  (-1, -5), [white, GRIS]),
        ("BACKGROUND",    (0, -4), (-1, -4), GRIS),
        ("FONTNAME",      (0, -4), (-1, -4), "Helvetica-Bold"),
        ("FONTSIZE",      (0, -4), (-1, -4), 9),
        ("BACKGROUND",    (0, -3), (-1, -3), AMAR),
        ("FONTNAME",      (0, -3), (-1, -3), "Helvetica-Bold"),
        ("FONTSIZE",      (0, -3), (-1, -3), 9),
        ("BACKGROUND",    (0, -2), (-1, -2), VERDE_C),
        ("FONTNAME",      (0, -2), (-1, -2), "Helvetica-Bold"),
        ("FONTSIZE",      (0, -2), (-1, -2), 10),
        ("TEXTCOLOR",     (2, -2), (2, -2),  VERDE_T),
        ("BACKGROUND",    (0, -1), (-1, -1), GRIS),
        ("TOPPADDING",    (0, 0),  (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 5),
        ("LEFTPADDING",   (0, 0),  (-1, -1), 6),
    ]))
    t2.wrapOn(c, w, h)
    t2.drawOn(c, 22, y - len(resumen_data) * 20 - 5)
    y_after = y - len(resumen_data) * 20 - 25

    # Bloque advertencia final
    c.setFillColor(AMAR)
    c.roundRect(22, y_after - 65, w - 44, 65, 5, fill=True, stroke=False)
    c.setStrokeColor(NARANJA)
    c.roundRect(22, y_after - 65, w - 44, 65, 5, fill=False, stroke=True)
    c.setFillColor(NARANJA)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(32, y_after - 14, "AVISOS IMPORTANTES PARA EL ASESOR FISCAL")
    c.setFont("Helvetica", 7.5)
    c.setFillColor(HexColor("#5A4A00"))
    avisos = [
        "• Reducción (cas. 0150): 60% para contratos anteriores al 26/05/2023. "
        "50% para posteriores (verificar si aplica 60%, 70% o 90%).",
        "• Amortización: calculada como MAX(precio compra total, valor catastral) "
        "× % construcción × 3% × % titularidad (Art. 14 RIRPF).",
        "• Los importes son proporcionales a los días de arrendamiento declarados "
        "en cada inmueble (casilla 0101).",
        "• Documento generado por Nolasco Capital. No sustituye asesoramiento "
        f"profesional. Ref: {ref} · {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    ]
    ya = y_after - 28
    for aviso in avisos:
        c.drawString(32, ya, aviso)
        ya -= 11

    c.save()
    buffer.seek(0)
    return buffer


# ────────────────────────────────────────────────────────────────
# 4. IMPORT EXCEL ASESOR — lee formato Álvaro y puebla Supabase
# ────────────────────────────────────────────────────────────────

def importar_excel_asesor(archivo_excel, user_id, upsert_inmueble_fn,
                           agregar_movimientos_fn, leer_inmuebles_fn):
    """
    Lee el Excel de Álvaro (formato real):
      - Hoja CONTABILIDAD: fila 1 = nombres inmuebles, col A = conceptos
      - Hojas individuales por inmueble: facturas de mantenimiento
      - Hoja Amortiz: datos catastrales y amortización
      - Usa upsert_inmueble → NUNCA borra, solo crea o actualiza
    """
    if not OPENPYXL_OK:
        return {"error": "Instala openpyxl: pip install openpyxl"}

    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    except Exception as e:
        return {"error": f"No se pudo leer el Excel: {e}"}

    if "CONTABILIDAD" not in wb.sheetnames:
        return {"error": "El Excel no tiene hoja CONTABILIDAD."}

    ws = wb["CONTABILIDAD"]
    filas_raw = list(ws.iter_rows(min_row=1, values_only=True))

    if not filas_raw:
        return {"error": "La hoja CONTABILIDAD está vacía."}

    # ── Fila 1 = nombres de inmuebles (col A = None, cols B... = nombres) ──
    cabecera = filas_raw[0]
    nombres_inmuebles = []
    col_indices = []
    for j, val in enumerate(cabecera):
        if j == 0:
            continue  # columna de conceptos
        if val and str(val).strip() not in ("TOTALES", "MENSUAL", "", "None"):
            nombres_inmuebles.append(str(val).strip())
            col_indices.append(j)

    if not nombres_inmuebles:
        return {"error": "No se encontraron inmuebles en la fila 1 de CONTABILIDAD."}

    # ── Mapeo concepto → campo Supabase ─────────────────────────
    MAPA = {
        "Referencia catastral":       "Ref_Catastral",
        "Superficie":                 "M2_Construidos",
        "Valor catastral":            "Valor_Catastral",
        "Titular":                    "Titular",
        "Nombre inquilino":           "Inquilino",
        "DNI inquilino":              "NIF_Inquilino",
        "Fecha contrato":             "Fecha_Inicio_Contrato",
        "Fecha adquisicion vivienda": "Fecha_Adquisicion",
        "INGRESOS":                   "_ingresos_anuales",
        "Mensuales":                  "Renta",
        "IBI":                        "IBI_Anual",
        "Comunidad":                  "_comunidad_anual",
        "Seguro vida":                "Seguro_Vida",
        "Seguro hogar":               "Seguro_Anual",
        "Amortizacion prestamo":      "Intereses_Hipoteca",
        "Ascensor":                   "Gasto_Ascensor",
        "Alarma":                     "_alarma",
        "Gastos Mantenimiento":       "_reparaciones",
        "IBI Cocheras":               "IBI_Cocheras",
        "Comunidad Cocheras":         "Comunidad_Cocheras",
        "Inmueble accesorio (garaje)":"Ref_Catastral_Cochera",
    }

    # ── Extraer datos por inmueble ───────────────────────────────
    datos = {nombre: {} for nombre in nombres_inmuebles}

    for fila in filas_raw[1:]:
        if not fila or not fila[0]:
            continue
        concepto = str(fila[0]).strip()
        if concepto not in MAPA:
            continue
        campo = MAPA[concepto]
        for idx, col_j in enumerate(col_indices):
            if col_j < len(fila):
                val = fila[col_j]
                if val is not None and str(val).strip() not in ("", "None"):
                    # Solo guardar si no tenemos ya un valor para este campo
                    if campo not in datos[nombres_inmuebles[idx]]:
                        datos[nombres_inmuebles[idx]][campo] = val

    # ── Leer hojas individuales de mantenimiento ─────────────────
    reparaciones_por_inmueble = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in ("CONTABILIDAD", "RESULTADOS", "Amortiz"):
            continue
        ws_inm = wb[sheet_name]
        filas_inm = list(ws_inm.iter_rows(min_row=1, values_only=True))
        if not filas_inm:
            continue
        # Buscar a qué inmueble pertenece esta hoja
        inm_match = None
        sheet_clean = sheet_name.upper().replace(" ", "").replace("Nº", "")
        for nombre in nombres_inmuebles:
            nom_clean = nombre.upper().replace(" ", "")
            if nom_clean in sheet_clean or sheet_clean in nom_clean:
                inm_match = nombre
                break
        if not inm_match:
            continue
        # Leer facturas (fila 1 = cabecera, resto = facturas)
        movs = []
        for fila in filas_inm[1:]:
            if not fila or not fila[0]:
                continue
            concepto_fac = str(fila[0]).strip()
            if concepto_fac.upper().startswith("TOTAL"):
                continue
            importe = fila[3] if len(fila) > 3 else None
            proveedor = fila[1] if len(fila) > 1 else ""
            if importe and float(importe) > 0:
                movs.append({
                    "concepto": concepto_fac,
                    "proveedor": str(proveedor) if proveedor else "",
                    "importe": float(importe),
                })
        if movs:
            reparaciones_por_inmueble[inm_match] = movs

    # ── Leer hoja Amortiz si existe ──────────────────────────────
    amortiz_por_inmueble = {}
    if "Amortiz" in wb.sheetnames:
        ws_am = wb["Amortiz"]
        filas_am = list(ws_am.iter_rows(min_row=1, values_only=True))
        for fila in filas_am:
            if not fila or not fila[0]:
                continue
            concepto_am = str(fila[0]).strip()
            for idx, nombre in enumerate(nombres_inmuebles):
                col_am = idx + 1
                if col_am < len(fila) and fila[col_am] is not None:
                    if nombre not in amortiz_por_inmueble:
                        amortiz_por_inmueble[nombre] = {}
                    if "precio compra" in concepto_am.lower() or "precio de compra" in concepto_am.lower():
                        amortiz_por_inmueble[nombre]["Precio_Compra"] = fila[col_am]
                    elif "valor catastral construccion" in concepto_am.lower():
                        amortiz_por_inmueble[nombre]["Valor_Real_Construccion"] = fila[col_am]
                    elif "% construccion" in concepto_am.lower() or "porcentaje construccion" in concepto_am.lower():
                        val = float(fila[col_am])
                        amortiz_por_inmueble[nombre]["Pct_Construccion"] = val if val <= 1 else val / 100
                    elif "amortizacion" in concepto_am.lower() and "%" not in concepto_am.lower():
                        amortiz_por_inmueble[nombre]["Amortizacion_Fiscal"] = fila[col_am]

    # ── Procesar cada inmueble ───────────────────────────────────
    import pandas as _pd
    hoy = datetime.now().strftime("%Y-%m-%d")
    creados      = []
    actualizados = []
    movimientos_nuevos = []

    for nombre in nombres_inmuebles:
        d = datos[nombre]
        registro = {"Nombre": nombre}

        # Campos directos
        for campo in [
            "Ref_Catastral", "M2_Construidos", "Valor_Catastral", "Titular",
            "Inquilino", "NIF_Inquilino", "Renta",
            "IBI_Anual", "Seguro_Vida", "Seguro_Anual", "Intereses_Hipoteca",
            "Gasto_Ascensor", "IBI_Cocheras", "Comunidad_Cocheras",
            "Ref_Catastral_Cochera",
        ]:
            if campo in d and d[campo] is not None:
                registro[campo] = d[campo]

        # Fecha contrato → string
        if "Fecha_Inicio_Contrato" in d:
            try:
                from datetime import datetime as _dt
                val = d["Fecha_Inicio_Contrato"]
                if hasattr(val, "strftime"):
                    registro["Fecha_Inicio_Contrato"] = val.strftime("%Y-%m-%d")
                else:
                    registro["Fecha_Inicio_Contrato"] = str(val)[:10]
            except:
                pass

        # Fecha adquisicion → string
        if "Fecha_Adquisicion" in d:
            try:
                val = d["Fecha_Adquisicion"]
                if hasattr(val, "strftime"):
                    registro["Fecha_Adquisicion"] = val.strftime("%Y-%m-%d")
                else:
                    registro["Fecha_Adquisicion"] = str(val)[:10]
            except:
                pass

        # Comunidad anual → mensual
        if "_comunidad_anual" in d:
            try:
                registro["Comunidad"] = round(float(d["_comunidad_anual"]) / 12, 2)
            except:
                pass

        # Alarma → sumarla a Gasto_Ascensor si existe
        if "_alarma" in d:
            try:
                alarma = float(d["_alarma"])
                actual = float(registro.get("Gasto_Ascensor", 0) or 0)
                registro["Gasto_Ascensor"] = round(actual + alarma, 2)
            except:
                pass

        # Datos de amortización
        if nombre in amortiz_por_inmueble:
            for k, v in amortiz_por_inmueble[nombre].items():
                registro[k] = v

        # Upsert en Supabase
        try:
            res = upsert_inmueble_fn(registro, user_id)
            if res.get("accion") == "creado":
                creados.append(nombre)
            else:
                actualizados.append(nombre)
        except Exception as e:
            actualizados.append(nombre)

        # Movimientos de reparación desde hojas individuales
        if nombre in reparaciones_por_inmueble:
            for rep in reparaciones_por_inmueble[nombre]:
                movimientos_nuevos.append({
                    "Fecha":       hoy,
                    "Apartamento": nombre,
                    "Concepto":    rep["concepto"],
                    "Categoría":   "Mantenimiento",
                    "Tipo":        "Gasto",
                    "Importe":     rep["importe"],
                    "Deducible":   "S",
                })

        # Movimiento de reparación desde CONTABILIDAD si no hay hoja individual
        elif "_reparaciones" in d:
            try:
                importe_rep = float(d["_reparaciones"])
                if importe_rep > 0:
                    movimientos_nuevos.append({
                        "Fecha":       hoy,
                        "Apartamento": nombre,
                        "Concepto":    "Gastos Mantenimiento (importado Excel)",
                        "Categoría":   "Mantenimiento",
                        "Tipo":        "Gasto",
                        "Importe":     importe_rep,
                        "Deducible":   "S",
                    })
            except:
                pass

    # Guardar movimientos
    if movimientos_nuevos:
        try:
            agregar_movimientos_fn(movimientos_nuevos, user_id)
        except:
            pass

    return {
        "ok":           True,
        "creados":      creados,
        "actualizados": actualizados,
        "movimientos":  len(movimientos_nuevos),
        "total":        len(creados) + len(actualizados),
    }
    if not OPENPYXL_OK:
        return {"error": "Instala openpyxl: pip install openpyxl"}

    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    except Exception as e:
        return {"error": f"No se pudo leer el Excel: {e}"}

    if "CONTABILIDAD" not in wb.sheetnames:
        return {"error": "El Excel no tiene hoja CONTABILIDAD. "
                         "Asegúrate de usar el Excel generado por Nolasco Capital."}

    ws = wb["CONTABILIDAD"]

    # ── Leer filas y columnas ────────────────────────────────────
    # Fila 3 = cabecera: col A = "CONCEPTO", cols B... = nombres inmuebles
    # Última columna = "TOTALES" — ignorar
    filas_raw = list(ws.iter_rows(min_row=1, values_only=True))

    # Encontrar fila de cabecera (la que tiene "CONCEPTO")
    fila_cab = None
    for i, fila in enumerate(filas_raw):
        if fila and str(fila[0]).strip() == "CONCEPTO":
            fila_cab = i
            break

    if fila_cab is None:
        return {"error": "No se encontró la fila de cabecera 'CONCEPTO' en la hoja CONTABILIDAD."}

    cabecera = filas_raw[fila_cab]
    # Nombres de inmuebles — desde col B hasta penúltima (última = TOTALES)
    nombres_inmuebles = []
    col_indices = []
    for j, val in enumerate(cabecera):
        if j == 0:
            continue  # columna CONCEPTO
        if val and str(val).strip() not in ("TOTALES", "", "None"):
            nombres_inmuebles.append(str(val).strip())
            col_indices.append(j)

    if not nombres_inmuebles:
        return {"error": "No se encontraron inmuebles en el Excel."}

    # ── Mapeo concepto → campo Supabase ─────────────────────────
    MAPA_CAMPOS = {
        "Ref. Catastral":            "Ref_Catastral",
        "Inquilino":                 "Inquilino",
        "NIF Inquilino":             "NIF_Inquilino",
        "Tipo Arrendamiento":        "Tipo_Arrendamiento",
        "Días arrendados (0101)":    "Dias_Arrendados_Anio",
        "Ingresos íntegros (0102)":  "_ingresos_anuales",   # especial → renta
        "Intereses hipoteca (0105)": "Intereses_Hipoteca",
        "Reparación/conserv. (0106)":"_reparaciones",        # especial → movimientos
        "IBI y tributos (0108)":     "IBI_Anual",
        "Comunidad+Seguros (0110)":  "_comunidad_seguros",   # especial → desglosar
        "Suministros (0111)":        "Servicios_Suministros",
        "Gastos jurídicos (0112)":   "Gastos_Juridicos",
        "Amortización 3% (0113)":    "Amortizacion_Fiscal",
        "Retenciones (0153)":        "Retenciones_IRPF",
    }

    # ── Extraer datos por inmueble ───────────────────────────────
    datos_por_inmueble = {nombre: {} for nombre in nombres_inmuebles}

    for fila in filas_raw[fila_cab + 1:]:
        if not fila or not fila[0]:
            continue
        concepto = str(fila[0]).strip()
        if concepto not in MAPA_CAMPOS:
            continue
        campo = MAPA_CAMPOS[concepto]
        for idx, col_j in enumerate(col_indices):
            if col_j < len(fila):
                val = fila[col_j]
                if val is not None and str(val).strip() not in ("", "None"):
                    datos_por_inmueble[nombres_inmuebles[idx]][campo] = val

    # ── Leer inmuebles actuales de Supabase ─────────────────────
    import pandas as _pd
    df_actual = leer_inmuebles_fn(user_id=user_id)
    nombres_actuales = {}
    if df_actual is not None and len(df_actual) > 0:
        for _, row in df_actual.iterrows():
            nombres_actuales[str(row.get("Nombre", "")).strip()] = row

    # ── Procesar cada inmueble ───────────────────────────────────
    creados     = []
    actualizados = []
    movimientos_nuevos = []
    hoy = datetime.now().strftime("%Y-%m-%d")

    for nombre, datos in datos_por_inmueble.items():
        # Construir registro base
        registro = {"Nombre": nombre}

        # Campos directos
        for campo_excel, campo_sup in [
            ("Ref_Catastral",        "Ref_Catastral"),
            ("Inquilino",            "Inquilino"),
            ("NIF_Inquilino",        "NIF_Inquilino"),
            ("Tipo_Arrendamiento",   "Tipo_Arrendamiento"),
            ("Dias_Arrendados_Anio", "Dias_Arrendados_Anio"),
            ("Intereses_Hipoteca",   "Intereses_Hipoteca"),
            ("IBI_Anual",            "IBI_Anual"),
            ("Servicios_Suministros","Servicios_Suministros"),
            ("Gastos_Juridicos",     "Gastos_Juridicos"),
            ("Amortizacion_Fiscal",  "Amortizacion_Fiscal"),
            ("Retenciones_IRPF",     "Retenciones_IRPF"),
        ]:
            if campo_excel in datos:
                registro[campo_sup] = datos[campo_excel]

        # Ingresos anuales → Renta mensual
        if "_ingresos_anuales" in datos:
            try:
                registro["Renta"] = round(float(datos["_ingresos_anuales"]) / 12, 2)
            except:
                pass

        # Comunidad+Seguros → Comunidad mensual (estimación)
        if "_comunidad_seguros" in datos:
            try:
                total_cs = float(datos["_comunidad_seguros"])
                # Aproximación: 60% comunidad, 40% seguro
                registro["Comunidad"] = round(total_cs * 0.60 / 12, 2)
                registro["Seguro_Anual"] = round(total_cs * 0.40, 2)
            except:
                pass

        # Si existe en Supabase → actualizar
        if nombre in nombres_actuales:
            fila_actual = nombres_actuales[nombre]
            # Mezclar: mantener datos actuales, añadir los que faltan
            for k, v in fila_actual.items():
                if k not in registro and v is not None and str(v) not in ("", "nan", "0"):
                    registro[k] = v
            actualizados.append(nombre)
        else:
            creados.append(nombre)

        # Upsert en Supabase — NUNCA borra, solo crea o actualiza este inmueble
        try:
            res = upsert_inmueble_fn(registro, user_id)
            if not res.get("ok"):
                pass  # Si falla uno, continuar con el resto
        except Exception as e:
            pass

        # Reparaciones → movimiento en diario contable
        if "_reparaciones" in datos:
            try:
                importe_rep = float(datos["_reparaciones"])
                if importe_rep > 0:
                    movimientos_nuevos.append({
                        "Fecha":      hoy,
                        "Apartamento": nombre,
                        "Concepto":   "Reparación y conservación (importado Excel)",
                        "Categoría":  "Mantenimiento",
                        "Tipo":       "Gasto",
                        "Importe":    importe_rep,
                        "Deducible":  "S",
                    })
            except:
                pass

    # Guardar movimientos de reparaciones
    if movimientos_nuevos:
        try:
            agregar_movimientos_fn(movimientos_nuevos, user_id)
        except:
            pass

    return {
        "ok":          True,
        "creados":     creados,
        "actualizados": actualizados,
        "movimientos": len(movimientos_nuevos),
        "total":       len(creados) + len(actualizados),
    }


# ────────────────────────────────────────────────────────────────
# 5. PDF INDIVIDUAL — un solo inmueble (sin import circular)
# ────────────────────────────────────────────────────────────────

def _generar_pdf_modelo100_individual(inmueble_data, modelo):
    """Genera PDF del Modelo 100 para un solo inmueble."""
    if not REPORTLAB_OK:
        return None
    from reportlab.pdfgen import canvas as rl_canvas2
    from reportlab.platypus import Table, TableStyle
    buffer = io.BytesIO()
    c = rl_canvas2.Canvas(buffer, pagesize=A4)
    w, h = A4
    azul_oscuro = HexColor("#0F2744")
    azul_acento = HexColor("#185FA5")
    verde       = HexColor("#1a7a40")
    naranja     = HexColor("#B45309")
    gris_claro  = HexColor("#F4F7FB")
    gris_borde  = HexColor("#D0DFF0")
    amarillo    = HexColor("#FFF9E6")
    ref      = f"NC-{datetime.now().strftime('%Y')}-{inmueble_data['Nombre'][:3].upper()}"
    tipo_arr = str(inmueble_data.get("Tipo_Arrendamiento", "Larga Duracion"))
    dias_arr = int(float(inmueble_data.get("Dias_Arrendados_Anio", 365) or 365))

    # Cabecera
    c.setFillColor(azul_oscuro)
    c.rect(0, h-100, w, 100, fill=True, stroke=False)
    c.setFillColor(azul_acento)
    c.roundRect(30, h-85, 55, 55, 6, fill=True, stroke=False)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 22); c.drawCentredString(57.5, h-65, "NC")
    c.setFont("Helvetica", 7); c.drawCentredString(57.5, h-77, "CAPITAL")
    c.setFont("Helvetica-Bold", 20); c.drawString(100, h-50, "Nolasco Capital")
    c.setFont("Helvetica", 9); c.drawString(100, h-65, "GRANADA  |  GESTION PATRIMONIAL INMOBILIARIA")
    c.setFont("Helvetica", 8)
    c.drawRightString(w-30, h-45, f"Ref: {ref}")
    c.drawRightString(w-30, h-57, f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
    c.drawRightString(w-30, h-69, f"Ejercicio: {datetime.now().year}")
    c.setStrokeColor(azul_acento); c.setLineWidth(3); c.line(0, h-103, w, h-103)

    # Título
    y = h-130
    c.setFillColor(azul_oscuro); c.setFont("Helvetica-Bold", 13)
    c.drawString(30, y, "Modelo 100 - Rendimientos del Capital Inmobiliario")
    c.setFont("Helvetica", 9); c.setFillColor(HexColor("#5A7A9A"))
    c.drawString(30, y-15, "Datos informativos para el asesor fiscal | Pendiente validacion profesional")

    # Datos inmueble
    y -= 45
    c.setFillColor(gris_claro); c.roundRect(25, y-70, w-50, 70, 6, fill=True, stroke=False)
    c.setStrokeColor(gris_borde); c.roundRect(25, y-70, w-50, 70, 6, fill=False, stroke=True)
    c.setFillColor(azul_oscuro); c.setFont("Helvetica-Bold", 10); c.drawString(35, y-14, "Datos del Inmueble")
    labels1 = [("Inmueble:", 35, str(inmueble_data["Nombre"])),
               ("Ref. Catastral:", 200, str(inmueble_data.get("Ref_Catastral","N/A"))),
               ("Titular:", 390, str(inmueble_data.get("Titular","N/A")))]
    labels2 = [("Modalidad:", 35, tipo_arr),
               ("NIF Inquilino:", 200, str(inmueble_data.get("NIF_Inquilino","N/A"))),
               ("Dias arrendados:", 390, f"{dias_arr} dias")]
    c.setFont("Helvetica", 8); c.setFillColor(HexColor("#5A7A9A"))
    for lbl, x, _ in labels1: c.drawString(x, y-30, lbl)
    for lbl, x, _ in labels2: c.drawString(x, y-48, lbl)
    c.setFillColor(azul_oscuro); c.setFont("Helvetica-Bold", 8)
    for lbl, x, val in labels1: c.drawString(x+60, y-30, val)
    for lbl, x, val in labels2: c.drawString(x+65, y-48, val)

    # Tabla casillas
    y -= 95
    c.setFillColor(azul_oscuro); c.setFont("Helvetica-Bold", 11); c.drawString(30, y, "Casillas del Modelo 100")
    c.setStrokeColor(azul_acento); c.setLineWidth(2); c.line(30, y-4, 230, y-4)
    y -= 22
    amort_detalle = modelo.get("0113_detalle", "")
    data_tabla = [
        ["Casilla", "Descripcion", "Importe (EUR)", "Nota"],
        ["0101", "Dias arrendado",                f"{modelo['0101']} dias",   ""],
        ["0102", "Ingresos integros",              f"{modelo['0102']:,.2f}",   ""],
        ["0105", "Intereses hipoteca",             f"{modelo['0105']:,.2f}",   "Solo intereses"],
        ["0106", "Reparacion y conservacion",      f"{modelo['0106']:,.2f}",   "Diario contable"],
        ["0108", "Tributos e IBI",                 f"{modelo['0108']:,.2f}",   ""],
        ["0110", "Comunidad + Seguros",            f"{modelo['0110']:,.2f}",   "Hogar+vida+comunidad"],
        ["0111", "Suministros",                    f"{modelo['0111']:,.2f}",   ""],
        ["0112", "Gastos juridicos",               f"{modelo['0112']:,.2f}",   ""],
        ["0113", "Amortizacion fiscal (3%)",       f"{modelo['0113']:,.2f}",   amort_detalle[:35] if amort_detalle else ""],
        ["0107", "TOTAL GASTOS DEDUCIBLES",        f"{modelo['0107']:,.2f}",   ""],
        ["0149", "RENDIMIENTO NETO",               f"{modelo['0149']:,.2f}",   ""],
        ["0150", f"Reduccion {modelo['reduccion_pct']}% (orient.)", f"-{modelo['0150']:,.2f}", "VALIDAR asesor"],
        ["0152", "RENDIMIENTO NETO REDUCIDO",      f"{modelo['0152']:,.2f}",   "VALIDAR asesor"],
        ["0153", "Retenciones practicadas",        f"{modelo['0153']:,.2f}",   ""],
    ]
    t = Table(data_tabla, colWidths=[58, 165, 90, 152])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  azul_oscuro),
        ("TEXTCOLOR",     (0,0),(-1,0),  white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  8),
        ("FONTNAME",      (0,1),(-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1),(-1,-1), 7.5),
        ("ALIGN",         (2,1),(2,-1),  "RIGHT"),
        ("FONTNAME",      (0,1),(0,-1),  "Helvetica-Bold"),
        ("GRID",          (0,0),(-1,-1), 0.4, gris_borde),
        ("LINEBELOW",     (0,0),(-1,0),  2, azul_acento),
        ("ROWBACKGROUNDS",(0,1),(-1,-5), [white, gris_claro]),
        ("BACKGROUND",    (0,-4),(-1,-4),HexColor("#F0F8FF")),
        ("FONTNAME",      (0,-4),(-1,-4),"Helvetica-Bold"),
        ("BACKGROUND",    (0,-3),(-1,-3),amarillo),
        ("FONTNAME",      (0,-3),(-1,-3),"Helvetica-Bold"),
        ("TEXTCOLOR",     (3,-3),(3,-3), naranja),
        ("BACKGROUND",    (0,-2),(-1,-2),HexColor("#FFF0DC")),
        ("FONTNAME",      (0,-2),(-1,-2),"Helvetica-Bold"),
        ("TEXTCOLOR",     (3,-2),(3,-2), naranja),
        ("BACKGROUND",    (0,-1),(-1,-1),gris_claro),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
    ]))
    t.wrapOn(c, w, h)
    t.drawOn(c, 25, y - len(data_tabla)*19 - 5)
    y_after = y - len(data_tabla)*19 - 20

    # Aviso
    y_av = y_after - 15
    c.setFillColor(amarillo); c.roundRect(25, y_av-55, w-50, 55, 5, fill=True, stroke=False)
    c.setStrokeColor(naranja); c.setLineWidth(1); c.roundRect(25, y_av-55, w-50, 55, 5, fill=False, stroke=True)
    c.setFillColor(naranja); c.setFont("Helvetica-Bold", 9)
    c.drawString(35, y_av-14, "IMPORTANTE: Documento informativo — pendiente validacion por asesor fiscal")
    c.setFont("Helvetica", 8); c.setFillColor(HexColor("#5A4A00"))
    c.drawString(35, y_av-28, "Reduccion 60% (cas. 0150): orientativa segun ingresos totales del contribuyente.")
    c.drawString(35, y_av-40, "Amortizacion: MAX(precio compra, catastral) x % construccion x 3%.")

    # Footer
    c.setFillColor(azul_oscuro); c.rect(0, 0, w, 30, fill=True, stroke=False)
    c.setFillColor(white); c.setFont("Helvetica", 7)
    c.drawString(30, 11, "Nolasco Capital  |  Granada  |  Gestion Patrimonial Inmobiliaria")
    c.drawRightString(w-30, 11, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.save()
    buffer.seek(0)
    return buffer


# ────────────────────────────────────────────────────────────────
# 6. RENDER STREAMLIT — sección Fiscalidad completa
# ────────────────────────────────────────────────────────────────

def render_seccion_fiscal(df_inm, df_mov, safe_float_fn, calcular_modelo_100_fn):
    """
    Reemplaza la sección Fiscalidad actual de app.py.
    Añade el resumen global y los botones de export.

    Pegar en app.py:
        elif menu == "Fiscalidad":
            from fiscal_export import render_seccion_fiscal
            render_seccion_fiscal(df_inm, df_mov, safe_float, calcular_modelo_100)
    """
    import streamlit as st

    st.markdown('<div class="nc-brand-header">Escudo Fiscal — Modelo 100 IRPF</div>',
                unsafe_allow_html=True)
    st.markdown('<div class="nc-brand-sub">Pre-relleno automático · Un inmueble o toda la cartera</div>',
                unsafe_allow_html=True)

    año_fiscal = st.selectbox("Ejercicio fiscal:", [2025, 2024, 2023],
                               index=0, key="año_fiscal_global")

    tab1, tab2 = st.tabs(["📋 Por inmueble", "📊 Resumen Global (todos sumados)"])

    # ── TAB 1: Por inmueble (igual que antes) ────────────────────
    with tab1:
        inmueble_sel = st.selectbox("Selecciona inmueble:",
                                     df_inm["Nombre"].tolist(),
                                     key="fiscal_inmueble_sel")
        fila = df_inm[df_inm["Nombre"] == inmueble_sel].iloc[0]
        modelo = calcular_modelo_100_fn(fila, df_mov, año_fiscal=año_fiscal)

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Ingresos Íntegros",   f"{modelo['0102']:,.0f} €", "Casilla 0102")
        k2.metric("Total Gastos",         f"{modelo['0107']:,.0f} €", "Deducibles")
        k3.metric("Rendimiento Neto",     f"{modelo['0149']:,.0f} €", "Casilla 0149")
        k4.metric("Base Imponible est.",  f"{modelo['0152']:,.0f} €",
                  f"Reducción {modelo['reduccion_pct']}%")

        st.markdown("---")
        st.markdown('<div class="nc-section-title">📋 Casillas Modelo 100</div>',
                    unsafe_allow_html=True)
        st.caption("Valores pre-rellenados desde tus datos · Confirmar con asesor fiscal")

        casillas = [
            {"Casilla": "0062-0075", "Descripción": "Identificación del inmueble",
             "Valor": modelo["0062_0075"], "Nota": ""},
            {"Casilla": "0076",  "Descripción": "Clave de uso",
             "Valor": modelo["0076"], "Nota": ""},
            {"Casilla": "0100",  "Descripción": "Reducción vivienda habitual",
             "Valor": modelo["0100"], "Nota": ""},
            {"Casilla": "0101",  "Descripción": "Días arrendado en el año",
             "Valor": f"{modelo['0101']} días", "Nota": ""},
            {"Casilla": "0102",  "Descripción": "Ingresos íntegros",
             "Valor": f"{modelo['0102']:,.2f} €", "Nota": ""},
            {"Casilla": "0105",  "Descripción": "Intereses hipoteca",
             "Valor": f"{modelo['0105']:,.2f} €", "Nota": "Solo intereses, no capital"},
            {"Casilla": "0106",  "Descripción": "Reparación y conservación",
             "Valor": f"{modelo['0106']:,.2f} €", "Nota": "Del diario contable"},
            {"Casilla": "0108",  "Descripción": "IBI y tributos",
             "Valor": f"{modelo['0108']:,.2f} €", "Nota": ""},
            {"Casilla": "0110",  "Descripción": "Comunidad + Seguros + Ascensor",
             "Valor": f"{modelo['0110']:,.2f} €", "Nota": "Seguro hogar + vida + comunidad"},
            {"Casilla": "0111",  "Descripción": "Servicios y suministros",
             "Valor": f"{modelo['0111']:,.2f} €", "Nota": ""},
            {"Casilla": "0112",  "Descripción": "Gastos jurídicos",
             "Valor": f"{modelo['0112']:,.2f} €", "Nota": ""},
            {"Casilla": "0113",  "Descripción": "Amortización fiscal (3%)",
             "Valor": f"{modelo['0113']:,.2f} €", "Nota": modelo.get("0113_detalle", "")},
            {"Casilla": "0107",  "Descripción": "TOTAL GASTOS DEDUCIBLES",
             "Valor": f"{modelo['0107']:,.2f} €", "Nota": ""},
            {"Casilla": "0149",  "Descripción": "RENDIMIENTO NETO",
             "Valor": f"{modelo['0149']:,.2f} €", "Nota": ""},
            {"Casilla": "0150",  "Descripción": f"Reducción {modelo['reduccion_pct']}% (orientativa)",
             "Valor": f"-{modelo['0150']:,.2f} €", "Nota": "⚠️ Validar con asesor"},
            {"Casilla": "0152",  "Descripción": "RENDIMIENTO NETO REDUCIDO",
             "Valor": f"{modelo['0152']:,.2f} €", "Nota": "⚠️ Pendiente validación"},
            {"Casilla": "0153",  "Descripción": "Retenciones practicadas",
             "Valor": f"{modelo['0153']:,.2f} €", "Nota": ""},
        ]
        st.dataframe(pd.DataFrame(casillas), use_container_width=True,
                     hide_index=True, height=620)

        if modelo.get("iva_aplicable"):
            st.warning("⚠️ IVA aplicable — tributa por Modelo 303, no en Modelo 100.")

        st.info(modelo.get("nota_reduccion", ""))

        st.markdown("---")
        if REPORTLAB_OK:
            if st.button("✅ Generar PDF este inmueble", type="primary",
                          use_container_width=True, key="pdf_inm"):
                pdf_buf = _generar_pdf_modelo100_individual(fila, modelo)
                if pdf_buf:
                    st.session_state["pdf_inm_buf"] = pdf_buf
                    st.success("✓ PDF generado")

            if "pdf_inm_buf" in st.session_state:
                st.download_button(
                    "📥 Descargar PDF inmueble",
                    data=st.session_state["pdf_inm_buf"],
                    file_name=f"Modelo100_{inmueble_sel.replace(' ','_')}_{año_fiscal}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

    # ── TAB 2: Resumen Global ─────────────────────────────────────
    with tab2:
        st.markdown("### 📊 Resumen de toda la cartera — IRPF " + str(año_fiscal))
        st.caption("Suma de todos los inmuebles · Formato Modelo 100 agregado")

        nombre_prop = st.text_input("Nombre propietario (para el documento):",
                                     value="Pedro Nolasco",
                                     key="nombre_prop_fiscal")
        nombre_ases = st.text_input("Nombre asesoría (aparece en el documento):",
                                     value="",
                                     placeholder="Ej: Quero Asesores Granada",
                                     key="nombre_ases_fiscal")

        with st.spinner("Calculando resumen global..."):
            filas, totales = calcular_resumen_global(
                df_inm, df_mov, safe_float_fn, calcular_modelo_100_fn,
                año_fiscal=año_fiscal
            )

        # KPIs globales
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Inmuebles",       totales["n_inmuebles"])
        k2.metric("Ingresos totales", f"{totales['ingresos']:,.0f}€",  "Casilla 0102")
        k3.metric("Total gastos",     f"{totales['total_gastos']:,.0f}€", "Deducibles")
        k4.metric("Rendimiento neto", f"{totales['rend_neto']:,.0f}€",  "Casilla 0149")
        k5.metric("Base imponible",   f"{totales['rend_final']:,.0f}€",
                  f"Red. est. {int(totales['reduccion_imp']/(totales['rend_neto'] or 1)*100)}%")

        st.markdown("---")

        # Tabla resumen por inmueble
        st.markdown("#### Por inmueble")
        df_tabla = pd.DataFrame([{
            "Inmueble":       f["inmueble"],
            "Ingresos (€)":   f"{f['ingresos']:,.0f}",
            "Gastos (€)":     f"{f['total_gastos']:,.0f}",
            "Amort. (€)":     f"{f['amortizacion']:,.0f}",
            "Rend. Neto (€)": f"{f['rend_neto']:,.0f}",
            "Reducción":      f"{f['reduccion_pct']}% ⚠️",
            "Base Imponible": f"{f['rend_final']:,.0f}",
        } for f in filas])
        st.dataframe(df_tabla, use_container_width=True, hide_index=True)

        # Totales en tabla visual
        st.markdown("#### Totales agregados — Modelo 100")
        totales_tabla = [
            {"Casilla": "0102", "Descripción": "Ingresos íntegros totales",
             "Total (€)": f"{totales['ingresos']:,.2f}", "Estado": "✓"},
            {"Casilla": "0105", "Descripción": "Intereses hipoteca",
             "Total (€)": f"{totales['intereses']:,.2f}", "Estado": "✓"},
            {"Casilla": "0106", "Descripción": "Reparación y conservación",
             "Total (€)": f"{totales['reparaciones']:,.2f}", "Estado": "✓"},
            {"Casilla": "0108", "Descripción": "IBI y tributos",
             "Total (€)": f"{totales['ibi']:,.2f}", "Estado": "✓"},
            {"Casilla": "0110", "Descripción": "Comunidad + Seguros",
             "Total (€)": f"{totales['comunidad_seguros']:,.2f}", "Estado": "✓"},
            {"Casilla": "0111", "Descripción": "Suministros",
             "Total (€)": f"{totales['suministros']:,.2f}", "Estado": "✓"},
            {"Casilla": "0112", "Descripción": "Gastos jurídicos",
             "Total (€)": f"{totales['gastos_juridicos']:,.2f}", "Estado": "✓"},
            {"Casilla": "0113", "Descripción": "Amortización fiscal 3%",
             "Total (€)": f"{totales['amortizacion']:,.2f}", "Estado": "✓"},
            {"Casilla": "0107", "Descripción": "TOTAL GASTOS DEDUCIBLES",
             "Total (€)": f"{totales['total_gastos']:,.2f}", "Estado": ""},
            {"Casilla": "0149", "Descripción": "RENDIMIENTO NETO",
             "Total (€)": f"{totales['rend_neto']:,.2f}", "Estado": ""},
            {"Casilla": "0150", "Descripción": "Reducción fiscal (orientativa)",
             "Total (€)": f"-{totales['reduccion_imp']:,.2f}", "Estado": "⚠️"},
            {"Casilla": "0152", "Descripción": "RENDIMIENTO NETO REDUCIDO",
             "Total (€)": f"{totales['rend_final']:,.2f}", "Estado": "⚠️"},
            {"Casilla": "0153", "Descripción": "Retenciones",
             "Total (€)": f"{totales['retenciones']:,.2f}", "Estado": "✓"},
        ]
        st.dataframe(pd.DataFrame(totales_tabla), use_container_width=True,
                     hide_index=True, height=490)

        st.warning(
            "⚠️ **Reducción fiscal orientativa.** "
            "Contratos anteriores al 26/05/2023 → 60%. "
            "Posteriores → 50% general (verificar si aplica 60%, 70% o 90%). "
            "Validar siempre con asesor fiscal."
        )

        st.markdown("---")
        st.markdown("#### 📥 Exportar para asesor fiscal")

        col1, col2 = st.columns(2)

        with col1:
            if st.button("📊 Generar Excel completo", type="primary",
                          use_container_width=True, key="gen_excel"):
                with st.spinner("Generando Excel..."):
                    excel_buf = generar_excel_asesor(
                        filas, totales,
                        nombre_propietario=nombre_prop,
                        nombre_asesoria=nombre_ases,
                        año_fiscal=año_fiscal
                    )
                if excel_buf:
                    st.session_state["excel_global"] = excel_buf
                    st.success("✓ Excel generado")
                else:
                    st.error("Instala openpyxl: pip install openpyxl")

            if "excel_global" in st.session_state:
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=st.session_state["excel_global"],
                    file_name=f"IRPF_{año_fiscal}_{nombre_prop.replace(' ','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        with col2:
            if st.button("📄 Generar PDF completo", type="primary",
                          use_container_width=True, key="gen_pdf_global"):
                with st.spinner("Generando PDF..."):
                    pdf_buf = generar_pdf_global(
                        filas, totales,
                        nombre_propietario=nombre_prop,
                        nombre_asesoria=nombre_ases,
                        año_fiscal=año_fiscal
                    )
                if pdf_buf:
                    st.session_state["pdf_global"] = pdf_buf
                    st.success("✓ PDF generado")
                else:
                    st.error("Instala reportlab: pip install reportlab")

            if "pdf_global" in st.session_state:
                st.download_button(
                    "⬇️ Descargar PDF",
                    data=st.session_state["pdf_global"],
                    file_name=f"IRPF_{año_fiscal}_{nombre_prop.replace(' ','_')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
